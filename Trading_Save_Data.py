import openpyxl

direction = "Data_MACD_RSI.xlsx"

def excel_store(end_time,type_trade,record_price,record_amount,currency_pair,row_y):
    wb = openpyxl.load_workbook(direction)
    ws = wb.worksheets[0]
    
    # Can be modified where to store data
    ws.cell(row=row_y, column=1).value = end_time# Date, column A 
    ws.cell(row=row_y, column=2).value = type_trade# Type column B
    ws.cell(row=row_y, column=3).value = record_price# Price column C 
    ws.cell(row=row_y, column=4).value = record_amount# Amount column D 
    ws.cell(row=row_y, column=5).value = currency_pair# Currency pair E 
    
    wb.save(direction)
